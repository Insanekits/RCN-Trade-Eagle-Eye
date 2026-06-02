#!/usr/bin/env python3
"""Sync Trade Eagle Eye dashboard data from the configured Excel workbook.

The script intentionally keeps dependencies light. Install ``openpyxl`` in the
runtime where the workbook is available, then run ``npm run sync``.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

DEFAULT_ENV_FILE = "sync.env"
DEFAULT_OUTPUT = "data/trade_eagle_eye.json"

FIELD_ALIASES = {
    "tradeId": ["trade id", "trade no", "contract no", "contract", "id"],
    "supplier": ["supplier", "shipper", "seller", "vendor"],
    "origin": ["origin", "country", "origin country"],
    "buyer": ["buyer", "customer", "entity"],
    "commodity": ["commodity", "product", "grade"],
    "quantityMt": ["quantity mt", "qty mt", "quantity", "mt", "volume"],
    "contractPriceUsd": ["contract price usd", "price usd", "price", "usd/mt", "rate"],
    "shipmentWindow": ["shipment window", "shipment", "eta", "etd", "month"],
    "status": ["status", "stage", "position"],
    "riskLevel": ["risk", "risk level", "alert"],
    "lastUpdated": ["last updated", "updated", "date"],
}


def parse_env(path: Path) -> Dict[str, str]:
    if not path.exists():
        raise FileNotFoundError(f"Environment file not found: {path}")

    values: Dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip().strip('"').strip("'")
        values[key.strip()] = value
    return values


def serialise(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def normalise_header(header: Any) -> str:
    return " ".join(str(header or "").strip().lower().replace("_", " ").split())


def canonicalise(row: Dict[str, Any]) -> Dict[str, Any]:
    normalised = {normalise_header(key): serialise(value) for key, value in row.items()}
    output: Dict[str, Any] = {}

    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalised and normalised[alias] not in (None, ""):
                output[canonical] = normalised[alias]
                break

    for key, value in normalised.items():
        if value not in (None, "") and key not in output:
            output[key] = value

    output.setdefault("status", "Monitor")
    output.setdefault("riskLevel", "Medium")
    return output


def read_workbook_rows(workbook_path: Path, sheet_name: str | None) -> Iterable[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install it where the workbook is available "
            "with `python -m pip install openpyxl`, then rerun `npm run sync`."
        ) from exc

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)

    headers = next(rows, None)
    if not headers:
        return []

    header_values = [str(header).strip() if header is not None else "" for header in headers]
    records: List[Dict[str, Any]] = []

    for row in rows:
        raw = {header_values[index]: value for index, value in enumerate(row) if index < len(header_values)}
        if any(value not in (None, "") for value in raw.values()):
            records.append(canonicalise(raw))

    return records


def build_payload(records: List[Dict[str, Any]], source: Path) -> Dict[str, Any]:
    return {
        "generatedAt": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "source": str(source),
        "records": records,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Trade Eagle Eye dashboard JSON from Excel.")
    parser.add_argument("--env", default=DEFAULT_ENV_FILE, help="Path to the sync.env file.")
    parser.add_argument("--validate-only", action="store_true", help="Validate configuration without reading Excel.")
    args = parser.parse_args()

    env_path = Path(args.env)
    env = parse_env(env_path)
    workbook_value = env.get("EXCEL_WORKBOOK_PATH", "").strip()
    output_path = Path(env.get("OUTPUT_JSON", DEFAULT_OUTPUT))
    sheet_name = env.get("SHEET_NAME", "").strip() or None

    if not workbook_value:
        print("EXCEL_WORKBOOK_PATH is required in sync.env", file=sys.stderr)
        return 1

    workbook_path = Path(os.path.expandvars(os.path.expanduser(workbook_value)))

    if args.validate_only:
        print(f"Configuration OK: {env_path}")
        print(f"Workbook path: {workbook_path}")
        print(f"Output path: {output_path}")
        if sheet_name:
            print(f"Sheet name: {sheet_name}")
        return 0

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    records = list(read_workbook_rows(workbook_path, sheet_name))
    payload = build_payload(records, workbook_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(records)} records to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
